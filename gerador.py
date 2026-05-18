"""
Gerador do Cronograma A3 do CEMFS.

Lê os dados das databases Atividades + Horários no Notion via API e produz
um arquivo .xlsx idêntico (em formato) ao QUADRO_GERAL original, com as
atividades atuais preenchidas nas células corretas.

Como funciona:
1. Carrega o template QUADRO_GERAL_TEMPLATE.xlsx (preserva fontes/bordas/page setup)
2. Limpa as células de atividade do template
3. Adiciona 2 colunas extras: P=Consultório Odontológico, Q=Hall Principal
4. Lê via API: Atividades (pra pegar periodicidade) + Horários (pra pegar slots)
5. Preenche cada slot no dia/sala/horário correto, colorindo conforme periodicidade
6. Salva como CRONOGRAMA_A3_CEMFS.xlsx

Requer:
- Variável de ambiente NOTION_TOKEN com o Internal Integration Secret
- Arquivo QUADRO_GERAL_TEMPLATE.xlsx na mesma pasta

Uso:
    NOTION_TOKEN=ntn_xxx python gerador.py
"""

import os
import sys
import json
from copy import copy
from datetime import datetime
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.cell.cell import MergedCell

# ============================================================
# CONFIGURAÇÃO — edite aqui se nomes ou IDs no Notion mudarem
# ============================================================

NOTION_TOKEN = os.environ.get('NOTION_TOKEN')
if not NOTION_TOKEN:
    print('ERRO: variável de ambiente NOTION_TOKEN não definida.', file=sys.stderr)
    sys.exit(1)

# IDs dos databases no Notion (parte da URL do banco, sem hífens)
DB_ATIVIDADES = '91887b98b6594ece8af7b1d23ae94d62'
DB_HORARIOS = '00582de716074af3ab9cf6d0dddc6ca8'

# Sala (Notion page ID, sem hífens) → coluna no xlsx (1-based)
SALA_TO_COL = {
    '3536a77010d8811eb06cca9c651067ef':  2,  # Sala 1 → B
    '3536a77010d881b896d0c5e053b7da1a':  3,  # Sala 2 → C
    '3536a77010d8815891cef8e1ceab4c18':  4,  # Sala 3 → D
    '3536a77010d881e68526c49564fdb27c':  5,  # Sala 4 → E
    '3536a77010d881ccbbc7cfd90fc08dd0':  6,  # Sala 5 → F
    '3536a77010d8810daae4e9e2ba46cf33':  7,  # Sala 6 → G
    '3536a77010d88142a7a2f2473c1af393':  8,  # Salão → H
    '3536a77010d88151b4c9d04012017745':  9,  # Costura → I
    '3536a77010d8818a9ee2fb63e2004704': 10,  # Refeitório → J
    '3536a77010d88100aef1eaa1d34f67b8': 11,  # Salão da Creche → K
    '3536a77010d881998a0bcdb50d8451e7': 12,  # Sala de Passe → L
    '3536a77010d881eab59efea7620542cf': 13,  # Pai João → M
    '3536a77010d881a4b22cda78aaf76fb1': 14,  # Atend. Fraterno → N
    '3536a77010d8812595eaef93283345ce': 15,  # H Refeit. → O
    '3536a77010d881b59317dc45f2d1cbb9': 16,  # Consultório Odont. → P
    '3536a77010d881c2bb74db947530e139': 17,  # Hall Principal → Q
}

DIA_TO_SHEET = {
    'Segunda': 'SEG', 'Terça': 'TER', 'Quarta': 'QUA',
    'Quinta': 'QUI', 'Sexta': 'SEX', 'Sábado': 'SAB', 'Domingo': 'DOM',
}

# Cor de fundo por periodicidade (idêntica ao QUADRO GERAL original)
PERIOD_COLOR = {
    'Semanal':   'C5E0B3',
    'Quinzenal': 'BDD6EE',
    'Mensal':    'F7CAAC',
    'Eventual':  'D883FF',
}

# Nome curto pra exibir no xlsx (espaço apertado, então abreviamos).
# Se um programa não está aqui, usa nome do programa em CAPS + corta no "—".
# Edite quando criar novos programas, ou quando quiser mudar o texto exibido.
NOMES_CURTOS = {
    'Consultas Odontológicas': 'CONSULTAS ODONT.',
    'Estudo do Livro dos Espíritos': 'ESTUDO LIVRO ESPÍRITOS',
    'Estudo do Livro dos Médiuns': 'ESTUDO LIVRO MÉDIUNS',
    'Estudo de O Céu e o Inferno': 'ESTUDO CÉU E INFERNO',
    'Reunião Mediúnica — Quarta': 'REUNIÃO MEDIÚNICA',
    'Ficha e Avaliação de Pacientes': 'FICHA / AVAL. PACIENTES',
    'Passe Tratamento Dr. Bitencourt': 'PASSE TRAT. DR. BITENCOURT',
    'Atendimento Dr. Bitencourt — Quarta manhã': 'DR. BITENCOURT',
    'Costura Fraterna': 'COSTURA',
    'Assistência à Comunidade (gestantes)': 'CURSO DE GESTANTES',
    'Reunião de Desobsessão': 'DESOBSESSÃO',
    'Estudo Doutrinário — Quarta 10:30': 'ESTUDO DOUTRINÁRIO',
    'Estudo Doutrinário — Quinta 18:30': 'ESTUDO DOUTRINÁRIO',
    'Estudo Doutrinário — Segunda 14:30': 'ESTUDO DOUTRINÁRIO',
    'Estudo Doutrinário — Sexta 08:00': 'ESTUDO DOUTRINÁRIO',
    'Estudo Doutrinário — Sexta 09:00': 'CULTO DO EVANGELHO',
    'Estudo Doutrinário — Quinta 07:00': 'ESTUDO DOUTRINÁRIO',
    'Estudo do Evangelho — Domingo 18:00': 'ESTUDO LIVRO ROTEIRO',
    'Evangelização dos Assistidos': 'EVANGELIZAÇÃO',
    'Recepção do Passe': 'RECEPÇÃO DO PASSE',
    'Irradiação — Quinta': 'IRRADIAÇÃO',
    'Irradiação — Segunda': 'IRRADIAÇÃO',
    'Palestra-Atendimento': 'PALESTRA-ATEND.',
    'Palestra — Quinta': 'PALESTRA',
    'Palestra — Segunda': 'PALESTRA',
    'Palestra — Terça (Dr. Bitencourt)': 'PALESTRA DR. BITENCOURT',
    'Palestra Pública': 'PALESTRA PÚBLICA',
    'Passe — Domingo 19:30': 'PASSE',
    'Passe Dr. Bitencourt — Quarta 08:00': 'PASSE DR. BITENCOURT',
    'Passe Dr. Bitencourt — Terça 16:00': 'PASSE DR. BITENCOURT',
    'Psicografia — Quinta': 'PSICOGRAFIA',
    'Psicografia — Terça': 'PSICOGRAFIA',
    'Atendimento Psicológico Infantil': 'PSICÓLOGO INFANTIL',
    'Atendimento Psicológico': 'PSICÓLOGO',
    'Reforço Escolar': 'REFORÇO ESCOLAR',
    'Reunião Administrativa': 'REUNIÃO ADM.',
    'Reunião Mediúnica — Sexta': 'REUNIÃO MEDIÚNICA',
    'Reunião Mediúnica — Quinta': 'REUNIÃO MEDIÚNICA',
    'Reunião Mediúnica — Ricardo': 'R. MEDIÚNICA RICARDO',
    'Reunião Mediúnica — Terça': 'REUNIÃO MEDIÚNICA',
    'Triagem de Doações': 'TRIAGEM DE DOAÇÕES',
    'Notícias Consoladoras': 'NOTÍCIAS CONSOL.',
}

# Linhas de dados no xlsx: 11 (07:00) até 41 (22:00), slots de 30min
DATA_ROW_START = 11
DATA_ROW_END = 41
ORIG_LAST_SALA_COL = 15  # O — última coluna do template original

NOTION_API = 'https://api.notion.com/v1'
HEADERS = {
    'Authorization': f'Bearer {NOTION_TOKEN}',
    'Notion-Version': '2022-06-28',
    'Content-Type': 'application/json',
}

# ============================================================
# Leitura do Notion via REST API
# ============================================================

def query_database(db_id):
    """Lista todas as páginas de um database, paginando."""
    results = []
    cursor = None
    while True:
        body = {}
        if cursor:
            body['start_cursor'] = cursor
        r = requests.post(f'{NOTION_API}/databases/{db_id}/query',
                          headers=HEADERS, json=body, timeout=30)
        if r.status_code != 200:
            print(f'ERRO API Notion: {r.status_code} {r.text}', file=sys.stderr)
            r.raise_for_status()
        data = r.json()
        results.extend(data.get('results', []))
        if not data.get('has_more'):
            break
        cursor = data.get('next_cursor')
    return results

# Helpers pra extrair valores de propriedades
def prop_title(prop):
    return ''.join(t.get('plain_text', '') for t in prop.get('title', []))

def prop_rich(prop):
    return ''.join(t.get('plain_text', '') for t in prop.get('rich_text', []))

def prop_select(prop):
    s = prop.get('select')
    return s.get('name', '') if s else ''

def prop_rel_ids(prop):
    return [r['id'].replace('-', '') for r in prop.get('relation', [])]

# ============================================================
# Geração do xlsx
# ============================================================

def time_to_slot(t):
    """'07:00' → 0, '07:30' → 1, ..., '22:00' → 30. Retorna None se inválido."""
    try:
        h, m = int(t[:2]), int(t[3:5])
        slot = (h - 7) * 2 + (1 if m >= 30 else 0)
        if slot < 0 or slot > 30:
            return None
        return slot
    except (ValueError, IndexError):
        return None

def clone_style(src, dst):
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.alignment = copy(src.alignment)
        dst.border = copy(src.border)
        dst.number_format = src.number_format

def safe_unmerge_at(ws, start_row, end_row, col):
    for mr in list(ws.merged_cells.ranges):
        if mr.min_col <= col <= mr.max_col:
            if not (mr.max_row < start_row or mr.min_row > end_row):
                ws.unmerge_cells(str(mr))

def nome_curto(nome_programa):
    """Retorna o nome curto pra exibir no xlsx."""
    if nome_programa in NOMES_CURTOS:
        return NOMES_CURTOS[nome_programa]
    base = nome_programa.split('—')[0].strip()
    return base.upper()[:30]

def build_workbook(horarios, template_path, out_path):
    """horarios: lista de dicts com dia, ini, fim, salas (ids), nome, period."""
    wb = load_workbook(template_path)

    for sheet_name in ['SEG','TER','QUA','QUI','SEX','SAB','DOM']:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # 1. Limpa células de atividade existentes (desmescla e zera)
        to_unmerge = []
        for mr in list(ws.merged_cells.ranges):
            if (mr.min_row >= DATA_ROW_START and
                mr.min_col >= 2 and mr.max_col <= ORIG_LAST_SALA_COL):
                to_unmerge.append(str(mr))
        for mr_str in to_unmerge:
            ws.unmerge_cells(mr_str)
        for r in range(DATA_ROW_START, DATA_ROW_END + 1):
            for c in range(2, ORIG_LAST_SALA_COL + 1):
                cell = ws.cell(row=r, column=c)
                cell.value = None
                cell.fill = PatternFill(fill_type=None)

        # 2. Estende merge do nome do dia (K2:O6 → K2:Q6)
        for mr in list(ws.merged_cells.ranges):
            if str(mr) == 'K2:O6':
                ws.unmerge_cells('K2:O6')
                break
        day_value = ws['K2'].value
        day_font, day_fill, day_align = copy(ws['K2'].font), copy(ws['K2'].fill), copy(ws['K2'].alignment)
        ws.merge_cells('K2:Q6')
        ws['K2'].value = day_value
        ws['K2'].font, ws['K2'].fill, ws['K2'].alignment = day_font, day_fill, day_align

        # 3. Configura colunas P (16) e Q (17)
        ws.column_dimensions['P'].width = 13.0
        ws.column_dimensions['Q'].width = 13.0
        for col in (16, 17):
            clone_style(ws.cell(row=8, column=15), ws.cell(row=8, column=col))
        clone_style(ws.cell(row=9, column=15), ws.cell(row=9, column=16))
        ws.cell(row=9, column=16, value='CONSULT. ODONT.')
        clone_style(ws.cell(row=9, column=15), ws.cell(row=9, column=17))
        ws.cell(row=9, column=17, value='HALL PRINC.')
        clone_style(ws.cell(row=10, column=15), ws.cell(row=10, column=16))
        clone_style(ws.cell(row=10, column=15), ws.cell(row=10, column=17))
        for r in range(DATA_ROW_START, DATA_ROW_END + 1):
            for col in (16, 17):
                c = ws.cell(row=r, column=col)
                clone_style(ws.cell(row=r, column=15), c)
                c.value = None
                c.fill = PatternFill(fill_type=None)

        # 4. Preenche atividades
        ref = ws.cell(row=DATA_ROW_START, column=4)
        conflitos = []
        for h in horarios:
            if DIA_TO_SHEET.get(h['dia']) != sheet_name:
                continue
            slot_ini = time_to_slot(h['ini'])
            slot_fim = time_to_slot(h['fim'])
            if slot_ini is None or slot_fim is None:
                continue
            n_slots = max(1, slot_fim - slot_ini)
            start_row = DATA_ROW_START + slot_ini
            end_row = min(DATA_ROW_END, start_row + n_slots - 1)

            for sala_id in h['salas']:
                col = SALA_TO_COL.get(sala_id)
                if not col:
                    continue
                safe_unmerge_at(ws, start_row, end_row, col)
                top = ws.cell(row=start_row, column=col)
                if isinstance(top, MergedCell):
                    conflitos.append(f"{sheet_name} {h['dia']} {h['ini']} col {col}")
                    continue
                if top.value:
                    top.value = f"{top.value}\n+ {h['nome']}"
                    conflitos.append(f"{sheet_name} {h['dia']} {h['ini']} col {col} (sobreposto)")
                else:
                    top.value = h['nome']
                fill = PatternFill('solid', start_color=PERIOD_COLOR.get(h['period'], 'C5E0B3'))
                top.fill = fill
                if not top.has_style:
                    clone_style(ref, top)
                for r in range(start_row, end_row + 1):
                    c = ws.cell(row=r, column=col)
                    if not isinstance(c, MergedCell):
                        c.fill = fill
                if end_row > start_row:
                    try:
                        ws.merge_cells(start_row=start_row, start_column=col,
                                       end_row=end_row, end_column=col)
                    except Exception:
                        pass
        if conflitos:
            print(f'  Conflitos em {sheet_name}: {len(conflitos)}')

    wb.save(out_path)

# ============================================================
# Main
# ============================================================

def main():
    template_path = 'QUADRO_GERAL_TEMPLATE.xlsx'
    out_path = 'CRONOGRAMA_A3_CEMFS.xlsx'

    if not os.path.exists(template_path):
        print(f'ERRO: arquivo {template_path} não encontrado.', file=sys.stderr)
        sys.exit(1)

    print(f'[{datetime.now():%Y-%m-%d %H:%M:%S}] Iniciando geração...')

    # Lê Atividades (pra mapear ID → nome + periodicidade)
    print('Lendo banco Atividades...')
    atividades_raw = query_database(DB_ATIVIDADES)
    atividades = {}
    for p in atividades_raw:
        pid = p['id'].replace('-', '')
        props = p['properties']
        atividades[pid] = {
            'nome': prop_title(props.get('Nome', {})),
            'periodicidade': prop_select(props.get('Periodicidade', {})) or 'Semanal',
        }
    print(f'  → {len(atividades)} programas encontrados')

    # Lê Horários
    print('Lendo banco Horários...')
    horarios_raw = query_database(DB_HORARIOS)
    horarios = []
    pulados = {'inativo': 0, 'modelo': 0, 'sem_dia_hora': 0}
    for p in horarios_raw:
        props = p['properties']
        nome_h = prop_title(props.get('Nome', {}))
        if 'MODELO' in nome_h.upper():
            pulados['modelo'] += 1
            continue
        status = prop_select(props.get('Status', {}))
        if status and status != 'Ativo':
            pulados['inativo'] += 1
            continue
        dia = prop_select(props.get('Dia', {}))
        ini = prop_rich(props.get('Início', {}))
        fim = prop_rich(props.get('Término', {}))
        if not dia or not ini or not fim:
            pulados['sem_dia_hora'] += 1
            continue

        atividade_ids = prop_rel_ids(props.get('Atividade', {}))
        sala_ids = prop_rel_ids(props.get('Salas', {}))

        nome_programa = ''
        period = 'Semanal'
        if atividade_ids and atividade_ids[0] in atividades:
            a = atividades[atividade_ids[0]]
            nome_programa = a['nome']
            period = a['periodicidade'] or 'Semanal'

        horarios.append({
            'dia': dia, 'ini': ini, 'fim': fim,
            'salas': sala_ids,
            'nome': nome_curto(nome_programa) if nome_programa else 'SEM PROGRAMA',
            'period': period,
        })
    print(f'  → {len(horarios)} horários ativos (pulados: {pulados})')

    # Gera xlsx
    print(f'Gerando {out_path}...')
    build_workbook(horarios, template_path, out_path)
    print(f'[{datetime.now():%Y-%m-%d %H:%M:%S}] Pronto.')

if __name__ == '__main__':
    main()
